r"""
Export Welling United Red OBDSFL workbook tables into clean dashboard JSON files.

Excel remains the source of truth. Attendance export contains football attendance/status data only;
monthly player fees are managed separately in the workbook and are not exported as attendance payment data.

Option A output: separate JSON files in ./data/
- players.json
- matches.json
- goals.json
- assists.json
- events.json
- attendance.json
- minutes.json

How to run from the dashboard repo folder:
    python export_welling_json.py

Optional workbook override:
    python export_welling_json.py --workbook "C:\Users\dansm\OneDrive\Documents\Dan\Football\Welling United Red OBDSFL 26-27.xlsx"

Or set an environment variable:
    setx WELLING_WORKBOOK_PATH "C:\Users\dansm\OneDrive\Documents\Dan\Football\Welling United Red OBDSFL 26-27.xlsx"

Requirements:
    pip install openpyxl

Recommended setup:
- Keep the Excel workbook in OneDrive as the editable source of truth.
- Keep only the generated JSON files in the GitHub dashboard repo.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

TEAM = "Welling United Red OBDSFL"
SEASON = "2026/27"
WORKBOOK_NAME = "Welling United Red OBDSFL 26-27.xlsx"
DATA_DIR = Path("data")
ENV_WORKBOOK_PATH = "WELLING_WORKBOOK_PATH"


def candidate_workbook_paths(script_root: Path) -> List[Path]:
    candidates: List[Path] = []
    env_path = os.environ.get(ENV_WORKBOOK_PATH)
    if env_path:
        candidates.append(Path(env_path).expanduser())
    candidates.append(script_root / WORKBOOK_NAME)
    candidates.append(Path.home() / "OneDrive" / "Documents" / "Dan" / "Football" / WORKBOOK_NAME)
    candidates.append(Path.home() / "OneDrive - Personal" / "Documents" / "Dan" / "Football" / WORKBOOK_NAME)
    candidates.append(Path.home() / "Documents" / "Dan" / "Football" / WORKBOOK_NAME)

    unique: List[Path] = []
    seen = set()
    for path in candidates:
        resolved_key = str(path)
        if resolved_key not in seen:
            unique.append(path)
            seen.add(resolved_key)
    return unique


def resolve_workbook_path(script_root: Path, workbook_arg: Optional[str] = None) -> Path:
    if workbook_arg:
        workbook_path = Path(workbook_arg).expanduser()
        if workbook_path.exists():
            return workbook_path
        raise FileNotFoundError(f"Workbook path from --workbook does not exist: {workbook_path}")

    candidates = candidate_workbook_paths(script_root)
    for path in candidates:
        if path.exists():
            return path

    candidate_text = "\n".join(f"- {path}" for path in candidates)
    raise FileNotFoundError(
        "Workbook not found. Checked these locations:\n"
        f"{candidate_text}\n\n"
        "Either move the workbook to one of those locations, run with --workbook, or set WELLING_WORKBOOK_PATH."
    )


def slugify(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.time().isoformat() == "00:00:00":
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def camel_key(header: Any) -> str:
    text = str(header or "").strip()
    text = text.replace("/", " ").replace("-", " ")
    parts = re.findall(r"[A-Za-z0-9]+", text)
    if not parts:
        return ""
    first = parts[0].lower()
    rest = [p[:1].upper() + p[1:].lower() for p in parts[1:]]
    key = first + "".join(rest)
    replacements = {
        "id": "id",
        "displayname": "displayName",
        "playerid": "playerId",
        "sessionid": "sessionId",
        "sessionkey": "sessionKey",
        "sessiondate": "sessionDate",
        "sessiontype": "sessionType",
        "submittedby": "submittedBy",
        "submittedat": "submittedAt",
        "matchid": "matchId",
        "matchdate": "matchDate",
        "recordtype": "recordType",
        "homeaway": "homeAway",
        "goalsfor": "goalsFor",
        "goalsagainst": "goalsAgainst",
    }
    return replacements.get(key.lower(), key)


def table_rows(workbook_path: Path, sheet_name: str, table_name: Optional[str] = None) -> List[Dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    if table_name is None:
        if not ws.tables:
            return []
        table_name = next(iter(ws.tables.keys()))

    if table_name not in ws.tables:
        return []

    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [clean_value(ws.cell(min_row, col).value) for col in range(min_col, max_col + 1)]
    keys = [camel_key(header) for header in headers]

    rows: List[Dict[str, Any]] = []
    for row_num in range(min_row + 1, max_row + 1):
        row: Dict[str, Any] = {}
        has_data = False
        for col_num, key in zip(range(min_col, max_col + 1), keys):
            value = clean_value(ws.cell(row_num, col_num).value)
            row[key] = value
            if value not in (None, ""):
                has_data = True
        if has_data:
            rows.append(row)
    return rows


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {path}")


def export_players(workbook_path: Path) -> List[Dict[str, Any]]:
    rows = table_rows(workbook_path, "Squad", "Squad")
    players = []
    for row in rows:
        player_id = row.get("id") or slugify(row.get("displayName") or row.get("name"))
        display_name = row.get("displayName") or row.get("name")
        if not player_id or not display_name or str(player_id).strip().lower() == "total":
            continue

        active_value = row.get("active")
        status_value = str(row.get("status") or "").strip().lower()
        active = bool(active_value) and status_value != "left"
        player = {"id": player_id, "displayName": display_name, "active": active}
        position = row.get("position")
        if position not in (None, ""):
            player["position"] = position
        players.append(player)
    return players


def export_matches(workbook_path: Path) -> List[Dict[str, Any]]:
    rows = table_rows(workbook_path, "Fixtures", "Fixtures")
    matches = []
    for row in rows:
        if row.get("opposition") in (None, "", 0, "0"):
            continue
        match_id = slugify(f"{row.get('date')}-{row.get('opposition')}")
        home_away = row.get("homeAway")
        venue = row.get("venue")
        matches.append({
            "id": match_id,
            "date": row.get("date"),
            "day": row.get("day"),
            "opposition": row.get("opposition"),
            "competition": row.get("competition"),
            "homeAway": home_away,
            "venue": venue or home_away,
            "postponed": bool(row.get("postponed")) if row.get("postponed") is not None else False,
            "goalsFor": row.get("goalsFor"),
            "goalsAgainst": row.get("goalsAgainst"),
            "result": row.get("result"),
        })
    return matches


def export_wide_player_stats(workbook_path: Path, sheet_name: str, output_key: str) -> List[Dict[str, Any]]:
    rows = table_rows(workbook_path, sheet_name, sheet_name)
    output = []
    ignored = {"date", "opposition", "count"}

    for row in rows:
        date_value = row.get("date")
        opposition = row.get("opposition")
        if opposition in (None, "", 0, "0"):
            continue

        match_id = slugify(f"{date_value}-{opposition}")
        players: Dict[str, Any] = {}
        for key, value in row.items():
            if key in ignored or value in (None, "", 0):
                continue
            player_key = str(key)
            if re.fullmatch(r"[a-z]+[A-Z][A-Za-z0-9]*", player_key):
                player_key = re.sub(r"([a-z0-9])([A-Z])", r"\1-\2", player_key).lower()
            players[player_key] = value

        output.append({"matchId": match_id, "date": date_value, "opposition": opposition, output_key: players})
    return output


def export_attendance(workbook_path: Path) -> Dict[str, Any]:
    rows = table_rows(workbook_path, "AttendanceRecords", "AttendanceRecords")
    player_names = {
        str(player.get("id")): str(player.get("displayName") or player.get("name") or player.get("id"))
        for player in export_players(workbook_path)
        if player.get("id")
    }
    sessions: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        session_key = row.get("sessionKey")
        player_id = row.get("playerId")
        display_name = player_names.get(str(player_id), row.get("displayName"))
        status = row.get("status")
        if not session_key or not player_id or not status:
            continue
        if session_key not in sessions:
            sessions[session_key] = {
                "sessionKey": session_key,
                "sessionId": row.get("sessionId"),
                "date": row.get("sessionDate"),
                "type": row.get("sessionType"),
                "venue": row.get("venue"),
                "submittedBy": row.get("submittedBy"),
                "submittedAt": row.get("submittedAt"),
                "records": [],
            }
        record = {
            "recordKey": row.get("recordKey"),
            "playerId": player_id,
            "displayName": display_name,
            "status": status,
            "source": row.get("source"),
        }
        record = {k: v for k, v in record.items() if v is not None}
        sessions[session_key]["records"].append(record)

    ordered_sessions = sorted(sessions.values(), key=lambda session: (session.get("date") or "", session.get("submittedAt") or ""))
    return {"team": TEAM, "season": SEASON, "sessions": ordered_sessions}


def _normal_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def export_minutes(workbook_path: Path) -> List[Dict[str, Any]]:
    """Export playing minutes directly from MatchdayRecords using raw headers.

    This deliberately avoids the generic camelCase row mapper because the workbook
    has carried both compact and spaced compatibility headers during migration.
    """
    wb = load_workbook(workbook_path, data_only=True)
    if "MatchdayRecords" not in wb.sheetnames:
        return []
    ws = wb["MatchdayRecords"]
    if "MatchdayRecords" not in ws.tables:
        return []

    table = ws.tables["MatchdayRecords"]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    raw_headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
    header_positions: Dict[str, int] = {}
    for offset, header in enumerate(raw_headers):
        key = _normal_header(header)
        if key and key not in header_positions:
            header_positions[key] = min_col + offset

    def col(*names: str) -> Optional[int]:
        for name in names:
            found = header_positions.get(_normal_header(name))
            if found is not None:
                return found
        return None

    columns = {
        "recordType": col("RecordType", "Record Type"),
        "sessionId": col("SessionId", "Session ID"),
        "matchId": col("MatchId", "Match ID"),
        "matchDate": col("MatchDate", "Match Date"),
        "opposition": col("Opposition"),
        "competition": col("Competition"),
        "playerId": col("PlayerId", "Player ID"),
        "displayName": col("DisplayName", "Display Name"),
        "value": col("Value"),
        "detail": col("Detail"),
        "submittedBy": col("SubmittedBy", "Submitted By"),
    }
    required = ("recordType", "playerId", "displayName", "value")
    if any(columns[name] is None for name in required):
        return []

    output: List[Dict[str, Any]] = []
    for row_num in range(min_row + 1, max_row + 1):
        def value(name: str) -> Any:
            c = columns.get(name)
            return clean_value(ws.cell(row_num, c).value) if c is not None else None

        if str(value("recordType") or "").strip().lower() != "minutes":
            continue
        player_id = value("playerId")
        display_name = value("displayName")
        if not player_id or not display_name:
            continue
        try:
            minutes = int(round(float(value("value") or 0)))
        except (TypeError, ValueError):
            minutes = 0
        output.append({
            "sessionId": value("sessionId"),
            "matchId": value("matchId"),
            "date": value("matchDate"),
            "opposition": value("opposition"),
            "competition": value("competition"),
            "playerId": player_id,
            "displayName": display_name,
            "minutes": minutes,
            "starter": str(value("detail") or "").strip().lower() == "starter",
            "submittedBy": value("submittedBy"),
        })
    return sorted(output, key=lambda item: (str(item.get("date") or ""), str(item.get("displayName") or "")))


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Welling dashboard JSON from the Excel workbook.")
    parser.add_argument("--workbook", help="Optional full path to the Excel workbook.")
    parser.add_argument("--data-dir", default=str(DATA_DIR), help="Output data folder. Default: data")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    workbook_path = resolve_workbook_path(root, args.workbook)
    data_dir = root / Path(args.data_dir)

    print(f"Using workbook: {workbook_path}")
    print(f"Writing JSON to: {data_dir}")

    write_json(data_dir / "players.json", export_players(workbook_path))
    write_json(data_dir / "matches.json", export_matches(workbook_path))
    write_json(data_dir / "goals.json", export_wide_player_stats(workbook_path, "Goals", "goals"))
    write_json(data_dir / "assists.json", export_wide_player_stats(workbook_path, "Assists", "assists"))
    write_json(data_dir / "events.json", export_wide_player_stats(workbook_path, "Events", "events"))
    write_json(data_dir / "attendance.json", export_attendance(workbook_path))
    write_json(data_dir / "minutes.json", export_minutes(workbook_path))

    print("Done. Dashboard JSON files updated.")


if __name__ == "__main__":
    main()
